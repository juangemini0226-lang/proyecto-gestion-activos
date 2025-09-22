from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction
from django.utils.text import slugify

from activos.models import (
    Activo,
    CategoriaActivo,
    ItemMantenible,
    Parte,
    Sistema,
    Subsistema,
    TipoUbicacion,
    Ubicacion,
)

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - handled during execution
    raise CommandError("Debe instalar la dependencia 'openpyxl' para usar este comando") from exc


Row = Dict[str, Optional[str]]


class Command(BaseCommand):
    help = (
        "Importa la jerarquía de activos (sistema → subsistema → componente → parte) "
        "a partir de un archivo Excel con las pestañas 'Activos', 'Sistemas',"
        " 'Subsistemas' y 'Partes'."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Ruta al archivo .xlsx que contiene la plantilla de importación",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])
        if not excel_path.exists():
            raise CommandError(f"No se encontró el archivo: {excel_path}")

        workbook = load_workbook(filename=str(excel_path), data_only=True)

        activos_rows = self._load_sheet(workbook, "Activos")
        sistemas_rows = self._load_sheet(workbook, "Sistemas")
        subsistemas_rows = self._load_sheet(workbook, "Subsistemas")
        partes_rows = self._load_sheet(workbook, "Partes")

        if not activos_rows:
            self.stdout.write(self.style.WARNING("La hoja 'Activos' está vacía o no existe."))

        resumen = self._importar(activos_rows, sistemas_rows, subsistemas_rows, partes_rows)

        self.stdout.write(self.style.SUCCESS("Importación finalizada."))
        for titulo, cantidad in resumen:
            self.stdout.write(f" - {titulo}: {cantidad}")

    def _importar(
        self,
        activos_rows: List[Row],
        sistemas_rows: List[Row],
        subsistemas_rows: List[Row],
        partes_rows: List[Row],
    ) -> List[Tuple[str, int]]:
        creados_activos = creados_sistemas = creados_subsistemas = 0
        creados_items = creados_partes = 0

        activos_cache: Dict[str, Activo] = {}
        sistemas_cache: Dict[Tuple[str, str], Sistema] = {}
        subsistemas_cache: Dict[Tuple[str, str, str], Subsistema] = {}
        items_cache: Dict[Tuple[str, str, str, str], ItemMantenible] = {}

        with transaction.atomic():
            for row in activos_rows:
                codigo = self._clean(row.get("CodigoActivo"))
                nombre = self._clean(row.get("NombreActivo"))
                if not codigo:
                    self.stderr.write("Fila de 'Activos' ignorada: falta 'CodigoActivo'.")
                    continue

                categoria_nombre = self._clean(row.get("Categoria"))
                categoria = None
                if categoria_nombre:
                    categoria, _ = CategoriaActivo.objects.get_or_create(nombre=categoria_nombre)

                ubicacion = self._ensure_ubicacion(
                    industria=self._clean(row.get("Industria")),
                    planta=self._clean(row.get("Planta")),
                    instalacion=self._clean(row.get("Instalacion")),
                )

                defaults = {
                    "numero_activo": codigo,
                    "nombre": nombre or codigo,
                    "categoria": categoria,
                    "ubicacion": ubicacion,
                }

                activo, created = Activo.objects.update_or_create(
                    codigo=codigo, defaults=defaults
                )
                if created:
                    creados_activos += 1
                else:
                    # update_or_create no actualiza FK si defaults contiene None
                    if activo.categoria != categoria:
                        activo.categoria = categoria
                    if activo.ubicacion != ubicacion:
                        activo.ubicacion = ubicacion
                    activo.nombre = defaults["nombre"]
                    activo.numero_activo = defaults["numero_activo"]
                    activo.save()

                activos_cache[codigo] = activo

            for row in sistemas_rows:
                codigo = self._clean(row.get("CodigoActivo"))
                nombre_sistema = self._clean(row.get("Sistema"))
                if not codigo or not nombre_sistema:
                    self.stderr.write(
                        "Fila de 'Sistemas' ignorada: se requieren 'CodigoActivo' y 'Sistema'."
                    )
                    continue

                activo = self._buscar_activo(codigo, activos_cache)
                if not activo:
                    self.stderr.write(
                        f"Fila de 'Sistemas' ignorada: no existe el activo con código {codigo}."
                    )
                    continue

                sistema, created = Sistema.objects.get_or_create(
                    activo=activo,
                    nombre=nombre_sistema,
                    defaults={
                        "tag": self._unique_tag(
                            Sistema,
                            self._build_tag(codigo, nombre_sistema),
                        )
                    },
                )
                if created:
                    creados_sistemas += 1
                sistemas_cache[(codigo, nombre_sistema)] = sistema

            for row in subsistemas_rows:
                codigo = self._clean(row.get("CodigoActivo"))
                nombre_sistema = self._clean(row.get("Sistema"))
                nombre_subsistema = self._clean(row.get("Subsistema"))
                if not codigo or not nombre_sistema or not nombre_subsistema:
                    self.stderr.write(
                        "Fila de 'Subsistemas' ignorada: faltan columnas requeridas."
                    )
                    continue

                sistema = sistemas_cache.get((codigo, nombre_sistema))
                if not sistema:
                    self.stderr.write(
                        "Fila de 'Subsistemas' ignorada: el sistema no existe; revise el orden de las hojas."
                    )
                    continue

                subsistema, created = Subsistema.objects.get_or_create(
                    sistema=sistema,
                    nombre=nombre_subsistema,
                    defaults={
                        "tag": self._unique_tag(
                            Subsistema,
                            self._build_tag(codigo, nombre_sistema, nombre_subsistema),
                        )
                    },
                )
                if created:
                    creados_subsistemas += 1
                subsistemas_cache[(codigo, nombre_sistema, nombre_subsistema)] = subsistema

            for row in partes_rows:
                codigo = self._clean(row.get("CodigoActivo"))
                nombre_sistema = self._clean(row.get("Sistema"))
                nombre_subsistema = self._clean(row.get("Subsistema"))
                nombre_item = self._clean(row.get("Componente"))
                nombre_parte = self._clean(row.get("Parte"))

                if not all([codigo, nombre_sistema, nombre_subsistema, nombre_item, nombre_parte]):
                    self.stderr.write(
                        "Fila de 'Partes' ignorada: faltan columnas requeridas (" \
                        "CodigoActivo, Sistema, Subsistema, Componente, Parte)."
                    )
                    continue

                subsistema = subsistemas_cache.get((codigo, nombre_sistema, nombre_subsistema))
                if not subsistema:
                    self.stderr.write(
                        "Fila de 'Partes' ignorada: el subsistema no existe; asegúrese de que esté definido en la hoja correspondiente."
                    )
                    continue

                item_clave = (codigo, nombre_sistema, nombre_subsistema, nombre_item)
                item = items_cache.get(item_clave)
                if not item:
                    item_tag = self._unique_tag(
                        ItemMantenible,
                        self._build_tag(codigo, nombre_sistema, nombre_subsistema, nombre_item),
                    )
                    item, created = ItemMantenible.objects.get_or_create(
                        subsistema=subsistema,
                        nombre=nombre_item,
                        defaults={"tag": item_tag},
                    )
                    if created:
                        creados_items += 1
                    items_cache[item_clave] = item

                tag_parte = self._clean(row.get("TAG"))
                if tag_parte:
                    tag_parte = tag_parte.strip()
                parte_tag = self._unique_tag(
                    Parte,
                    tag_parte or self._build_tag(
                        codigo, nombre_sistema, nombre_subsistema, nombre_item, nombre_parte
                    ),
                )

                parte_defaults = {"tag": parte_tag, "nombre": nombre_parte}
                try:
                    parte, created = Parte.objects.get_or_create(
                        item=item,
                        nombre=nombre_parte,
                        defaults=parte_defaults,
                    )
                except IntegrityError:
                    self.stderr.write(
                        f"No se pudo crear la parte '{nombre_parte}' (tag duplicado)."
                    )
                    continue

                if created:
                    creados_partes += 1
                else:
                    if parte.tag != parte_tag:
                        parte.tag = parte_tag
                        parte.save(update_fields=["tag"])

        return [
            ("Activos nuevos", creados_activos),
            ("Sistemas nuevos", creados_sistemas),
            ("Subsistemas nuevos", creados_subsistemas),
            ("Componentes nuevos", creados_items),
            ("Partes nuevas", creados_partes),
        ]

    def _clean(self, value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
        else:
            value = str(value).strip()
        return value or None

    def _load_sheet(self, workbook, sheet_name: str) -> List[Row]:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        header = [self._clean(col) or "" for col in rows[0]]
        data_rows: List[Row] = []
        for raw in rows[1:]:
            if all(value is None or str(value).strip() == "" for value in raw):
                continue
            data_rows.append({header[idx]: self._clean(value) for idx, value in enumerate(raw)})
        return data_rows

    def _build_tag(self, *parts: Optional[str]) -> str:
        fragments: List[str] = []
        for index, part in enumerate(parts):
            if not part:
                continue
            slug = slugify(part, allow_unicode=False)
            slug = slug.replace("-", "_").upper()
            if not slug:
                slug = f"SEG{index}"
            fragments.append(slug)
        return "_".join(fragments) or "TAG"

    def _unique_tag(self, model, base: str) -> str:
        base = re.sub(r"[^A-Z0-9_]+", "_", base.upper()).strip("_") or "TAG"
        base = re.sub(r"_+", "_", base)
        if len(base) > 100:
            base = base[:100]
        candidate = base
        suffix = 1
        while model.objects.filter(tag=candidate).exists():
            extra = f"_{suffix}"
            suffix += 1
            candidate = f"{base[: 100 - len(extra)]}{extra}"
        return candidate

    def _buscar_activo(self, codigo: str, cache: Dict[str, Activo]) -> Optional[Activo]:
        if codigo in cache:
            return cache[codigo]
        try:
            activo = Activo.objects.get(codigo=codigo)
        except Activo.DoesNotExist:
            return None
        cache[codigo] = activo
        return activo

    def _ensure_ubicacion(
        self,
        *,
        industria: Optional[str],
        planta: Optional[str],
        instalacion: Optional[str],
    ) -> Optional[Ubicacion]:
        ubicacion = None
        if industria:
            ubicacion, _ = Ubicacion.objects.get_or_create(
                nombre=industria,
                padre=None,
                defaults={"tipo": TipoUbicacion.INDUSTRIA},
            )
            if ubicacion.tipo != TipoUbicacion.INDUSTRIA:
                ubicacion.tipo = TipoUbicacion.INDUSTRIA
                ubicacion.save(update_fields=["tipo"])

        if planta:
            ubicacion, _ = Ubicacion.objects.get_or_create(
                nombre=planta,
                padre=ubicacion,
                defaults={"tipo": TipoUbicacion.PLANTA},
            )
            if ubicacion.tipo != TipoUbicacion.PLANTA:
                ubicacion.tipo = TipoUbicacion.PLANTA
                ubicacion.save(update_fields=["tipo"])

        if instalacion:
            ubicacion, _ = Ubicacion.objects.get_or_create(
                nombre=instalacion,
                padre=ubicacion,
                defaults={"tipo": TipoUbicacion.PROCESO},
            )
            if ubicacion.tipo != TipoUbicacion.PROCESO:
                ubicacion.tipo = TipoUbicacion.PROCESO
                ubicacion.save(update_fields=["tipo"])

        return ubicacion
