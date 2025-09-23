"""Herramientas para importar la taxonomía de un activo desde Excel."""

import io
import unicodedata
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable

from django.db import transaction

from openpyxl import load_workbook

from .models import ItemMantenible, Parte, Sistema, Subsistema

__all__ = [
    "TaxonomiaImportError",
    "TaxonomiaImportSummary",
    "TaxonomiaImporter",
]


class TaxonomiaImportError(Exception):
    """Error fatal durante la importación."""


class TaxonomiaRowError(Exception):
    """Error localizado en una fila concreta del archivo."""


@dataclass(slots=True)
class TaxonomiaImportSummary:
    """Resultados de la importación de taxonomía."""

    sistemas_creados: int = 0
    sistemas_actualizados: int = 0
    subsistemas_creados: int = 0
    subsistemas_actualizados: int = 0
    items_creados: int = 0
    items_actualizados: int = 0
    partes_creadas: int = 0
    partes_actualizados: int = 0
    errores: list[str] = field(default_factory=list)

    @property
    def total_creados(self) -> int:
        return (
            self.sistemas_creados
            + self.subsistemas_creados
            + self.items_creados
            + self.partes_creadas
        )

    @property
    def total_actualizados(self) -> int:
        return (
            self.sistemas_actualizados
            + self.subsistemas_actualizados
            + self.items_actualizados
            + self.partes_actualizados
        )

    def build_message(self) -> str:
        """Devuelve un resumen legible del resultado."""

        partes = []
        if self.total_creados:
            partes.append(f"{self.total_creados} elementos creados")
        if self.total_actualizados:
            partes.append(f"{self.total_actualizados} actualizados")
        if not partes:
            partes.append("No se realizaron cambios")
        return "; ".join(partes)


class TaxonomiaImporter:
    """Carga jerarquías de taxonomía (ISO 14224) desde un archivo Excel."""

    HEADER_ALIASES: Dict[str, set[str]] = {
        "sistema_tag": {
            "sistema_tag",
            "tag_sistema",
            "tag_del_sistema",
            "tag sistema",
        },
        "sistema_codigo": {
            "sistema_codigo",
            "codigo_sistema",
            "codigo sistema",
            "cod_sistema",
        },
        "sistema_nombre": {
            "sistema_nombre",
            "nombre_sistema",
            "nombre sistema",
            "sistema",
        },
        "subsistema_tag": {
            "subsistema_tag",
            "tag_subsistema",
            "tag subsistema",
        },
        "subsistema_codigo": {
            "subsistema_codigo",
            "codigo_subsistema",
            "codigo subsistema",
            "cod_subsistema",
        },
        "subsistema_nombre": {
            "subsistema_nombre",
            "nombre_subsistema",
            "nombre subsistema",
            "subsistema",
        },
        "item_tag": {
            "item_tag",
            "tag_item",
            "tag item",
            "tag_item_mantenible",
        },
        "item_codigo": {
            "item_codigo",
            "codigo_item",
            "codigo item",
            "cod_item",
        },
        "item_nombre": {
            "item_nombre",
            "nombre_item",
            "nombre item",
            "item",
            "item_mantenible",
        },
        "parte_tag": {
            "parte_tag",
            "tag_parte",
            "tag parte",
        },
        "parte_codigo": {
            "parte_codigo",
            "codigo_parte",
            "codigo parte",
            "cod_parte",
        },
        "parte_nombre": {
            "parte_nombre",
            "nombre_parte",
            "nombre parte",
            "parte",
        },
    }

    def __init__(self, *, activo, archivo, limpiar: bool = False):
        if activo is None:
            raise ValueError("Debe especificarse el activo a actualizar.")
        if archivo is None:
            raise ValueError("Debe proporcionarse un archivo de taxonomía.")

        self.activo = activo
        self.archivo = archivo
        self.limpiar = limpiar

    # ------------------------------------------------------------------
    # API pública
    # ------------------------------------------------------------------
    def importar(self) -> TaxonomiaImportSummary:
        """Importa la jerarquía completa y devuelve un resumen."""

        wb = self._open_workbook()
        sheet = wb.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise TaxonomiaImportError("El archivo no contiene encabezados.")

        header_map = self._build_header_map(header_row)
        if "sistema_tag" not in header_map and "sistema_nombre" not in header_map:
            raise TaxonomiaImportError(
                "El archivo debe incluir al menos una columna de identificación del sistema."
            )

        summary = TaxonomiaImportSummary()

        with transaction.atomic():
            if self.limpiar:
                # El borrado se revierte automáticamente si algo falla.
                self.activo.sistemas.all().delete()

            caches: Dict[str, Dict[str, Any]] = {
                "sistemas": {},
                "subsistemas": {},
                "items": {},
                "partes": {},
            }

            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                data = self._extract_row_data(row, header_map)
                if not any(data.values()):
                    continue
                try:
                    self._import_row(data, caches, summary)
                except TaxonomiaRowError as exc:
                    summary.errores.append(f"Fila {row_idx}: {exc}")

        return summary

    # ------------------------------------------------------------------
    # Procesamiento fila por fila
    # ------------------------------------------------------------------
    def _import_row(
        self,
        data: Dict[str, str],
        caches: Dict[str, Dict[str, Any]],
        summary: TaxonomiaImportSummary,
    ) -> None:
        sistema = self._ensure_sistema(data, caches, summary)
        subsistema = self._ensure_subsistema(data, sistema, caches, summary)
        item = self._ensure_item(data, subsistema, caches, summary)
        self._ensure_parte(data, item, caches, summary)

    # ------------------------------------------------------------------
    # Creación / actualización de cada nivel
    # ------------------------------------------------------------------
    def _ensure_sistema(
        self,
        data: Dict[str, str],
        caches: Dict[str, Dict[str, Any]],
        summary: TaxonomiaImportSummary,
    ) -> Sistema:
        tag = data.get("sistema_tag", "")
        nombre = data.get("sistema_nombre", "")
        codigo = data.get("sistema_codigo", "")

        if not (tag or nombre or codigo):
            raise TaxonomiaRowError("La fila no contiene información del sistema.")
        if not tag:
            raise TaxonomiaRowError("Falta el tag del sistema.")

        if tag in caches["sistemas"]:
            sistema = caches["sistemas"][tag]
            if self._update_model(sistema, {"nombre": nombre, "codigo": codigo}):
                summary.sistemas_actualizados += 1
            return sistema

        try:
            sistema = Sistema.objects.get(tag=tag)
            if sistema.activo_id != self.activo.id:
                raise TaxonomiaRowError(
                    f"El sistema '{tag}' pertenece al activo {sistema.activo.codigo}."
                )
            if self._update_model(sistema, {"nombre": nombre, "codigo": codigo}):
                summary.sistemas_actualizados += 1
        except Sistema.DoesNotExist:
            sistema = Sistema.objects.create(
                activo=self.activo,
                tag=tag,
                nombre=nombre or "",
                codigo=codigo or "",
            )
            summary.sistemas_creados += 1

        caches["sistemas"][tag] = sistema
        return sistema

    def _ensure_subsistema(
        self,
        data: Dict[str, str],
        sistema: Sistema,
        caches: Dict[str, Dict[str, Any]],
        summary: TaxonomiaImportSummary,
    ) -> Subsistema | None:
        tag = data.get("subsistema_tag", "")
        nombre = data.get("subsistema_nombre", "")
        codigo = data.get("subsistema_codigo", "")

        if not (tag or nombre or codigo):
            return None
        if not tag:
            raise TaxonomiaRowError("Falta el tag del subsistema.")

        if tag in caches["subsistemas"]:
            subsistema = caches["subsistemas"][tag]
            if subsistema.sistema_id != sistema.id:
                raise TaxonomiaRowError(
                    f"El subsistema '{tag}' pertenece al sistema {subsistema.sistema.tag}."
                )
            if self._update_model(subsistema, {"nombre": nombre, "codigo": codigo}):
                summary.subsistemas_actualizados += 1
            return subsistema

        try:
            subsistema = Subsistema.objects.get(tag=tag)
            if subsistema.sistema_id != sistema.id:
                raise TaxonomiaRowError(
                    f"El subsistema '{tag}' pertenece al sistema {subsistema.sistema.tag}."
                )
            if self._update_model(subsistema, {"nombre": nombre, "codigo": codigo}):
                summary.subsistemas_actualizados += 1
        except Subsistema.DoesNotExist:
            subsistema = Subsistema.objects.create(
                sistema=sistema,
                tag=tag,
                nombre=nombre or "",
                codigo=codigo or "",
            )
            summary.subsistemas_creados += 1

        caches["subsistemas"][tag] = subsistema
        return subsistema

    def _ensure_item(
        self,
        data: Dict[str, str],
        subsistema: Subsistema | None,
        caches: Dict[str, Dict[str, Any]],
        summary: TaxonomiaImportSummary,
    ) -> ItemMantenible | None:
        tag = data.get("item_tag", "")
        nombre = data.get("item_nombre", "")
        codigo = data.get("item_codigo", "")

        if not (tag or nombre or codigo):
            return None
        if subsistema is None:
            raise TaxonomiaRowError("Hay información de ítem pero falta el subsistema.")
        if not tag:
            raise TaxonomiaRowError("Falta el tag del ítem mantenible.")

        if tag in caches["items"]:
            item = caches["items"][tag]
            if item.subsistema_id != subsistema.id:
                raise TaxonomiaRowError(
                    f"El ítem '{tag}' pertenece al subsistema {item.subsistema.tag}."
                )
            if self._update_model(item, {"nombre": nombre, "codigo": codigo}):
                summary.items_actualizados += 1
            return item

        try:
            item = ItemMantenible.objects.get(tag=tag)
            if item.subsistema_id != subsistema.id:
                raise TaxonomiaRowError(
                    f"El ítem '{tag}' pertenece al subsistema {item.subsistema.tag}."
                )
            if self._update_model(item, {"nombre": nombre, "codigo": codigo}):
                summary.items_actualizados += 1
        except ItemMantenible.DoesNotExist:
            item = ItemMantenible.objects.create(
                subsistema=subsistema,
                tag=tag,
                nombre=nombre or "",
                codigo=codigo or "",
            )
            summary.items_creados += 1

        caches["items"][tag] = item
        return item

    def _ensure_parte(
        self,
        data: Dict[str, str],
        item: ItemMantenible | None,
        caches: Dict[str, Dict[str, Any]],
        summary: TaxonomiaImportSummary,
    ) -> Parte | None:
        tag = data.get("parte_tag", "")
        nombre = data.get("parte_nombre", "")
        codigo = data.get("parte_codigo", "")

        if not (tag or nombre or codigo):
            return None
        if item is None:
            raise TaxonomiaRowError("Hay información de parte pero falta el ítem mantenible.")
        if not tag:
            raise TaxonomiaRowError("Falta el tag de la parte.")

        if tag in caches["partes"]:
            parte = caches["partes"][tag]
            if parte.item_id != item.id:
                raise TaxonomiaRowError(
                    f"La parte '{tag}' pertenece al ítem {parte.item.tag}."
                )
            if self._update_model(parte, {"nombre": nombre, "codigo": codigo}):
                summary.partes_actualizados += 1
            return parte

        try:
            parte = Parte.objects.get(tag=tag)
            if parte.item_id != item.id:
                raise TaxonomiaRowError(
                    f"La parte '{tag}' pertenece al ítem {parte.item.tag}."
                )
            if self._update_model(parte, {"nombre": nombre, "codigo": codigo}):
                summary.partes_actualizados += 1
        except Parte.DoesNotExist:
            parte = Parte.objects.create(
                item=item,
                tag=tag,
                nombre=nombre or "",
                codigo=codigo or "",
            )
            summary.partes_creadas += 1

        caches["partes"][tag] = parte
        return parte

    # ------------------------------------------------------------------
    # Utilidades
    # ------------------------------------------------------------------
    def _open_workbook(self):
        if hasattr(self.archivo, "read"):
            content = self.archivo.read()
            try:
                self.archivo.seek(0)
            except Exception:  # pragma: no cover - algunos streams no son seekables
                pass
        else:
            content = self.archivo

        if not content:
            raise TaxonomiaImportError("El archivo de taxonomía está vacío.")

        if isinstance(content, str):
            content = content.encode()

        return load_workbook(io.BytesIO(content), data_only=True)

    def _build_header_map(self, headers: Iterable[Any]) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for idx, header in enumerate(headers):
            key = self._normalize_header(header)
            if not key:
                continue
            for field, aliases in self.HEADER_ALIASES.items():
                if key in aliases and field not in mapping:
                    mapping[field] = idx
                    break
        return mapping

    def _extract_row_data(self, row: Iterable[Any], mapping: Dict[str, int]) -> Dict[str, str]:
        data: Dict[str, str] = {}
        row_values = list(row)
        for field, idx in mapping.items():
            value = row_values[idx] if idx < len(row_values) else None
            data[field] = self._clean_cell(value)
        return data

    def _clean_cell(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    def _normalize_header(self, header: Any) -> str:
        if header is None:
            return ""
        text = str(header)
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.category(ch).startswith("M"))
        text = text.lower()
        for ch in ("-", "/"):
            text = text.replace(ch, " ")
        return "_".join(text.split())

    def _update_model(self, instance, values: Dict[str, str]) -> bool:
        update_fields: list[str] = []
        for field, value in values.items():
            if value is None:
                continue
            if isinstance(value, str):
                value = value.strip()
            if not value:
                continue
            if getattr(instance, field) != value:
                setattr(instance, field, value)
                update_fields.append(field)
        if update_fields:
            instance.save(update_fields=update_fields)
            return True
        return False
