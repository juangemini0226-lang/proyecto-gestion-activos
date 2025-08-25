# horometro/services/importer.py
import io
import re
import unicodedata
import datetime as dt
from dataclasses import dataclass
from typing import Dict, Optional, Iterable, Tuple
from decimal import Decimal

from django.db import transaction
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from openpyxl import load_workbook
from activos.models import Activo
from horometro.models import LecturaHorometro
from .alerts import sync_alert_for_reading  # sincroniza/crea/cierra alertas


# ===== Utilidades =====
COLUMN_MAP = {
    "codigo": ["ACTIVO","CODIGO","CÓDIGO","CODIGO ACTIVO","CÓDIGO ACTIVO","ID","ASSET_CODE","COD ACTIVO","COD. ACTIVO","COD"],
    "numero_activo": ["NUMERO ACTIVO","NÚMERO ACTIVO","NRO ACTIVO","NRO. ACTIVO","Nº ACTIVO","N° ACTIVO","NO ACTIVO","NUM ACTIVO","ASSET_NO"],
    "lectura": ["MEDIDOR","HORAS","CICLOS","LECTURA"],  # lectura semanal (o “Ciclos actuales Sx” en ancho)
    "anio": ["AÑO","ANIO","YEAR"],
    "semana": ["SEMANA","WEEK"],
    "fecha": ["FECHA","DATE"],
    "nombre": ["NOMBRE ACTIVO","NOMBRE","ASSET_NAME"],
    # NO semanales:
    "ciclos_oracle": ["CICLOS ORACLE", "CICLOS_ORACLE", "ORACLE"],
}

def _norm(s: str) -> str:
    return (str(s) if s is not None else "").strip().upper()

def _norm_ascii(s: str) -> str:
    """Mayúsculas + sin acentos + espacios simples."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.category(ch).startswith("M"))
    s = " ".join(s.upper().split())
    return s

def _find_col(name: str, headers: Iterable[str]) -> Optional[str]:
    wanted = set(map(_norm, COLUMN_MAP.get(name, [])))
    for h in headers:
        if _norm(h) in wanted:
            return h
    return None

def _match_week_header(h) -> Optional[int]:
    """Detecta cabeceras de semana: 'SEMANA 1', 'CICLOS ACTUALES S1', 'S1'."""
    if not h:
        return None
    txt = str(h).strip()
    m = re.match(r"^\s*SEMANA\s+(\d{1,2})\s*$", txt, flags=re.IGNORECASE)
    if m: return int(m.group(1))
    m = re.match(r"^\s*CICLOS\s+ACTUALES\s*S\s*(\d{1,2})\s*$", txt, flags=re.IGNORECASE)
    if m: return int(m.group(1))
    m = re.match(r"^\s*S\s*(\d{1,2})\s*$", txt, flags=re.IGNORECASE)
    if m: return int(m.group(1))
    return None

def _find_week_col(headers: Iterable[str], base_labels: Iterable[str], week: int) -> Optional[int]:
    """
    Busca cabecera tipo '<base> S<week>' ignorando acentos/mayus.
    Ej.: base='ÚLTIMO PREVENTIVO', week=1 => 'ULTIMO PREVENTIVO S1'
    """
    wants = {_norm_ascii(f"{b} S{int(week)}") for b in base_labels}
    for i, h in enumerate(headers):
        if _norm_ascii(h) in wants:
            return i
    return None

def _to_decimal(val) -> Optional[Decimal]:
    if val in (None, "", "#N/A"):
        return None
    try:
        return Decimal(str(val).replace(",", "."))
    except Exception:
        return None

def _compute_year_week(fecha) -> Optional[Tuple[int, int]]:
    if fecha is None: return None
    if isinstance(fecha, str):
        fecha = dt.date.fromisoformat(fecha.strip())
    if hasattr(fecha, "isocalendar"):
        iso = fecha.isocalendar()
        # Compatibilidad py3.8/3.12
        year = getattr(iso, "year", iso[0])
        week = getattr(iso, "week", iso[1])
        return int(year), int(week)
    return None

def _clean_sheet_name(nombre_hoja: Optional[str]) -> Optional[str]:
    if nombre_hoja is None: return None
    s = str(nombre_hoja).strip()
    return s if s else None


@dataclass
class ImportResult:
    creados: int = 0
    actualizados: int = 0
    errores: int = 0
    match_log: str = ""
    @property
    def resumen(self) -> str:
        return f"{self.creados} creados, {self.actualizados} actualizados, {self.errores} errores"


# ===== Importador principal =====
def importar_excel(
    archivo,
    nombre_hoja: Optional[str] = None,
    dry_run: bool = True,
    usuario=None,
    anio_fijo: Optional[int] = None,
    semana_fija: Optional[int] = None,
    generar_alertas: bool = True,   # controla la sincronización de alertas
) -> Dict:
    """
    Carga lecturas del horómetro.

    Si se pasa anio_fijo/semana_fija, se usan para TODAS las filas (semana por semana).
    Además importa: ciclos_oracle, ciclo_ultimo_preventivo (Sx), ciclos_desde_ultimo_preventivo (Sx).

    Parámetros:
      - dry_run: si True, simula y no guarda nada.
      - generar_alertas: si True, sincroniza alertas por fila y al final con la ÚLTIMA lectura por activo.
    """
    nombre_hoja = _clean_sheet_name(nombre_hoja)

    # Leer archivo a memoria
    content_bytes = archivo.read() if hasattr(archivo, "read") else archivo.getvalue()
    memfile = io.BytesIO(content_bytes)

    # Aún no guardamos el archivo: lo haremos la primera vez que tengamos (año, semana)
    saved_file_path = None
    orig_name = getattr(archivo, "name", "horometro.xlsx")

    # Abrir workbook
    wb = load_workbook(memfile, data_only=True, read_only=True)
    ws = wb[nombre_hoja] if nombre_hoja else wb.active

    # ---- 1) Headers ----
    header_row_idx = None
    headers = None
    rows_probe = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
    for idx, row in enumerate(rows_probe, start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(_norm(h) in map(_norm, COLUMN_MAP["codigo"]) for h in cells) \
           or any(_norm(h) in map(_norm, COLUMN_MAP["numero_activo"]) for h in cells) \
           or any(_norm(h) in map(_norm, COLUMN_MAP["nombre"]) for h in cells):
            header_row_idx, headers = idx, cells
            break
    if not headers:
        for idx, row in enumerate(rows_probe, start=1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(_match_week_header(h) for h in cells):
                header_row_idx, headers = idx, cells
                break
    if not headers:
        header_row_idx = 1
        headers = [str(c).strip() if c is not None else "" for c in ws[1]]

    idxmap = {h: i for i, h in enumerate(headers)}

    # ---- 2) ¿Formato ancho? ----
    semana_cols = []
    for h in headers:
        wk = _match_week_header(h)
        if wk:
            semana_cols.append((h, wk))
    is_wide = len(semana_cols) > 0

    # índices opcionales (no por semana)
    idx_ciclos_oracle = None
    col_c_oracle = _find_col("ciclos_oracle", headers)
    if col_c_oracle:
        idx_ciclos_oracle = idxmap.get(col_c_oracle)

    # bases de columnas semanales extra
    BASES_SEMANA = {
        "ultimo_prev": ["ULTIMO PREVENTIVO", "ÚLTIMO PREVENTIVO"],
        "delta_prev": ["CICLOS DESDE EL ULTIMO PREVENTIVO", "CICLOS DESDE EL ÚLTIMO PREVENTIVO"],
        # la lectura en ancho se toma de semana_cols (ya detectada)
    }

    res = ImportResult()
    log_lines, err_lines = [], []
    default_year = dt.date.today().year
    data_rows = ws.iter_rows(min_row=header_row_idx + 1, values_only=True)

    # Activos tocados en este import (para sincronización final de última semana)
    touched_activos = set()

    # Guardar una sola copia cuando tengamos y/w definitivos
    def ensure_saved_file(y, w):
        nonlocal saved_file_path
        if dry_run or saved_file_path or not content_bytes:
            return
        folder = f"horometro/imports/{int(y)}/{int(w):02d}/"
        saved_file_path = default_storage.save(folder + orig_name, ContentFile(content_bytes))

    # ==== 3A) Formato ANCHO ====
    if is_wide:
        col_codigo = _find_col("codigo", headers)
        col_num = _find_col("numero_activo", headers)
        col_nombre = _find_col("nombre", headers)
        col_anio = _find_col("anio", headers)  # opcional

        if not (col_codigo or col_num or col_nombre):
            raise ValueError("No se encontró columna de activo (ACTIVO/CODIGO, NUMERO ACTIVO o NOMBRE ACTIVO).")

        @transaction.atomic
        def _do_wide():
            nonlocal res
            row_idx = header_row_idx
            for row in data_rows:
                row_idx += 1
                try:
                    r = list(row)
                    codigo = str(r[idxmap[col_codigo]]).strip() if col_codigo and idxmap.get(col_codigo) is not None else ""
                    numero = str(r[idxmap[col_num]]).strip() if col_num and idxmap.get(col_num) is not None else ""
                    nombre = str(r[idxmap[col_nombre]]).strip() if col_nombre and idxmap.get(col_nombre) is not None else ""

                    # Buscar activo
                    qs = Activo.objects.all()
                    if codigo:
                        qs = qs.filter(codigo__iexact=codigo)
                    elif numero:
                        qs = qs.filter(numero_activo__iexact=numero)
                    elif nombre:
                        qs = qs.filter(nombre__iexact=nombre)
                    activo = qs.first()
                    if not activo:
                        res.errores += 1
                        err_lines.append(f"Fila {row_idx}: Activo no encontrado (codigo='{codigo}', numero='{numero}', nombre='{nombre}')")
                        continue

                    # Año base
                    if col_anio and idxmap.get(col_anio) is not None:
                        try:
                            anio_val = int(str(r[idxmap[col_anio]]).strip())
                        except Exception:
                            anio_val = default_year
                    else:
                        anio_val = default_year
                    if anio_fijo:
                        anio_val = int(anio_fijo)

                    # NO semanal
                    c_oracle = _to_decimal(r[idx_ciclos_oracle]) if idx_ciclos_oracle is not None and idx_ciclos_oracle < len(r) else None

                    # Por cada Semana N con dato -> si semana_fija, solo esa
                    for col_name, semana_num in semana_cols:
                        if semana_fija and int(semana_num) != int(semana_fija):
                            continue
                        ci = idxmap.get(col_name)
                        val = r[ci] if ci is not None and ci < len(r) else None
                        lectura = _to_decimal(val)
                        if lectura is None:
                            # sin lectura no grabamos fila
                            continue

                        semana_final = int(semana_fija) if semana_fija else int(semana_num)
                        ensure_saved_file(anio_val, semana_final)

                        # Semanales extras S<semana_final>
                        idx_up = _find_week_col(headers, BASES_SEMANA["ultimo_prev"], semana_final)
                        idx_dp = _find_week_col(headers, BASES_SEMANA["delta_prev"],  semana_final)

                        ciclo_up   = _to_decimal(r[idx_up]) if idx_up is not None and idx_up < len(r) else None
                        delta_prev = _to_decimal(r[idx_dp]) if idx_dp is not None and idx_dp < len(r) else None

                        defaults = {
                            "lectura": lectura,
                            "creado_por": usuario if not dry_run else None,
                            "fila_excel": row_idx,
                            "ciclos_oracle": c_oracle,
                            "ciclo_ultimo_preventivo": ciclo_up,
                            "ciclos_desde_ultimo_preventivo": delta_prev,
                        }
                        if not dry_run:
                            obj, created = LecturaHorometro.objects.update_or_create(
                                activo=activo, anio=anio_val, semana=semana_final, defaults=defaults
                            )
                            if saved_file_path and not obj.fuente_archivo:
                                obj.fuente_archivo.name = saved_file_path
                                obj.save(update_fields=["fuente_archivo"])
                            # sincronizar alerta (si se solicitó)
                            if generar_alertas:
                                sync_alert_for_reading(obj)

                            touched_activos.add(activo.pk)
                            res.creados += int(created)
                            res.actualizados += int(not created)

                        log_lines.append(f"{activo.codigo or activo.numero_activo} -> {anio_val}-W{semana_final:02d} = {lectura}")

                except Exception as e:
                    res.errores += 1
                    err_lines.append(f"Fila {row_idx}: {e}")

            # Sincroniza alerta con la ÚLTIMA lectura por activo tocado (evita “ruido histórico”)
            if not dry_run and generar_alertas and touched_activos:
                for aid in touched_activos:
                    latest = (LecturaHorometro.objects
                              .filter(activo_id=aid)
                              .order_by('-anio', '-semana')
                              .first())
                    if latest:
                        sync_alert_for_reading(latest)

            if dry_run:
                transaction.set_rollback(True)

        _do_wide()
        return {"resumen": res.resumen, "match_log": "\n".join(log_lines[:5000]), "errores": "\n".join(err_lines[:5000])}

    # ==== 3B) Formato ESTRECHO ====
    col_codigo = _find_col("codigo", headers)
    col_num = _find_col("numero_activo", headers)
    col_nombre = _find_col("nombre", headers)
    col_lect = _find_col("lectura", headers)
    col_anio = _find_col("anio", headers)
    col_sem = _find_col("semana", headers)
    col_fecha = _find_col("fecha", headers)

    if not col_lect or not (col_codigo or col_num or col_nombre) or not (col_sem or col_fecha):
        raise ValueError("Encabezados requeridos no encontrados: activo(código/número/nombre), lectura y (semana o fecha)")

    @transaction.atomic
    def _do_narrow():
        nonlocal res
        row_idx = header_row_idx
        for row in data_rows:
            row_idx += 1
            try:
                r = list(row)
                codigo = str(r[idxmap[col_codigo]]).strip() if col_codigo else ""
                numero = str(r[idxmap[col_num]]).strip() if col_num else ""
                nombre = str(r[idxmap[col_nombre]]).strip() if col_nombre else ""
                lectura = _to_decimal(r[idxmap[col_lect]])

                if lectura is None:
                    # sin lectura no grabamos fila
                    err_lines.append(f"Fila {row_idx}: Sin valor de lectura.")
                    continue

                # Año y semana detectados
                if col_anio and col_sem:
                    anio = int(str(r[idxmap[col_anio]]).strip() or dt.date.today().year)
                    semana = int(str(r[idxmap[col_sem]]).strip())
                elif col_fecha:
                    y_w = _compute_year_week(r[idxmap[col_fecha]])
                    if not y_w:
                        raise ValueError("No se pudo convertir FECHA a año/semana")
                    anio, semana = y_w
                else:
                    anio = dt.date.today().year
                    semana = int(str(r[idxmap[col_sem]]).strip())

                # Sobrescribir por selección del formulario (semana por semana)
                if anio_fijo:   anio = int(anio_fijo)
                if semana_fija: semana = int(semana_fija)

                # Buscar activo
                qs = Activo.objects.all()
                if codigo:
                    qs = qs.filter(codigo__iexact=codigo)
                elif numero:
                    qs = qs.filter(numero_activo__iexact=numero)
                elif nombre:
                    qs = qs.filter(nombre__iexact=nombre)
                activo = qs.first()
                if not activo:
                    res.errores += 1
                    err_lines.append(f"Fila {row_idx}: Activo no encontrado (codigo='{codigo}', numero='{numero}', nombre='{nombre}')")
                    continue

                ensure_saved_file(anio, semana)

                # NO semanal
                c_oracle = _to_decimal(r[idx_ciclos_oracle]) if idx_ciclos_oracle is not None and idx_ciclos_oracle < len(r) else None

                # Semanales extras para esta semana
                idx_up = _find_week_col(headers, BASES_SEMANA["ultimo_prev"], semana)
                idx_dp = _find_week_col(headers, BASES_SEMANA["delta_prev"],  semana)

                ciclo_up   = _to_decimal(r[idx_up]) if idx_up is not None and idx_up < len(r) else None
                delta_prev = _to_decimal(r[idx_dp]) if idx_dp is not None and idx_dp < len(r) else None

                defaults = {
                    "lectura": lectura,
                    "creado_por": usuario if not dry_run else None,
                    "fila_excel": row_idx,
                    "ciclos_oracle": c_oracle,
                    "ciclo_ultimo_preventivo": ciclo_up,
                    "ciclos_desde_ultimo_preventivo": delta_prev,
                }
                if not dry_run:
                    obj, created = LecturaHorometro.objects.update_or_create(
                        activo=activo, anio=anio, semana=semana, defaults=defaults
                    )
                    if saved_file_path and not obj.fuente_archivo:
                        obj.fuente_archivo.name = saved_file_path
                        obj.save(update_fields=["fuente_archivo"])
                    # sincronizar alerta (si se solicitó)
                    if generar_alertas:
                        sync_alert_for_reading(obj)

                    touched_activos.add(activo.pk)
                    res.creados += int(created)
                    res.actualizados += int(not created)

                log_lines.append(f"{activo.codigo or activo.numero_activo} -> {anio}-W{semana:02d} = {lectura}")

            except Exception as e:
                res.errores += 1
                err_lines.append(f"Fila {row_idx}: {e}")

        # Sincroniza alerta con la ÚLTIMA lectura por activo tocado
        if not dry_run and generar_alertas and touched_activos:
            for aid in touched_activos:
                latest = (LecturaHorometro.objects
                          .filter(activo_id=aid)
                          .order_by('-anio', '-semana')
                          .first())
                if latest:
                    sync_alert_for_reading(latest)

        if dry_run:
            transaction.set_rollback(True)

    _do_narrow()
    return {
        "resumen": res.resumen,
        "match_log": "\n".join(log_lines[:5000]),
        "errores": "\n".join(err_lines[:5000]),
    }
